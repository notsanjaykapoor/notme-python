import typing

import openpyxl


def stream_xls(file: str) -> typing.Generator[tuple[int, dict], None, None]:
    wb = openpyxl.load_workbook(file)
    ws = wb[wb.sheetnames[0]]

    headers = None

    for row in ws.iter_rows():
        values = [cell.value for cell in row]

        # check if row looks like a header line
        if not headers and _sheet_header(values=values):
            headers = values
            continue

        if not headers:
            # no headers yet, so keep parsing
            continue

        object = dict(zip(headers, values))
        object.pop(None, None)

        yield 1, object


def _sheet_header(values: list[str]) -> bool:
    if not values[0] or not values[1]:
        # row is empty
        return False

    return True
